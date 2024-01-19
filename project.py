import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import time
from openpyxl import Workbook, load_workbook 

service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=option)

dati=load_workbook('dati_.xlsx')
kanali=dati['YT_kanali']
balso=dati['NBA_Balsosana']
#'''
lastRow = kanali.max_row
for rinda in range(2,lastRow+1):
    saite=(kanali['d' + str(rinda)].value)
    pedejais = (kanali['b' + str(rinda)].value)
    url = str(saite)
    driver.get(url)
    time.sleep(2)
    cookies_no = driver.find_element(By.CLASS_NAME,"VfPpkd-LgbsSe VfPpkd-LgbsSe-OWXEXe-k8QpJ VfPpkd-LgbsSe-OWXEXe-dgl2Hf nCP5yc AjY5Oe DuMIQc iMLaPd")
    cookies_no.click
    time.sleep(2)
    latest = driver.find_element(By.ID,"video-title-link")
    saite = latest.get_attribute("href")
    jaunais = saite
    if(pedejais == jaunais):
        kanali['c'+str(rinda)].value= "tas pats"
    else:
        kanali['c'+str(rinda)].value= "JAUNS"
        kanali['b'+str(rinda)].value= jaunais
#'''
lastRow = balso.max_row
# Līdz : Sun, 21 Jan 2024, 06:59
url = "https://vote.nba.com/en/search.html"
email = "temp_adress@email.lv"
pasw = "********"
driver.get(url)
time.sleep(2)
cookies_no = driver.find_element(By.ID,"onetrust-reject-all-handler")
cookies_no.click
time.sleep(2)
for rinda in range(1,lastRow+1):
    balss = (balso['a' + str(rinda)].value)
    search = driver.find_element(By.CLASS_NAME,"input_JjHAN")
    search.clear
    search.send_keys(balss)
    button = driver.find_element(By.CLASS_NAME,"button_2STL- largeSubmit_agTaX")
    button.click
review = driver.find_element(By.CLASS_NAME,"button_zhKS-")
review.click
rev_sub = driver.find_element(By.CLASS_NAME,"button_zhKS-")
rev_sub.click
email_input=driver.find_element(By.ID,"email")
pasw_input=driver.find_element(By.ID,"password")
email_input.send_keys(email)
pasw_input.send_keys(pasw)
login = driver.find_element(By.CLASS_NAME,"Button_button__L2wUb SignIn_signInButton__CjbA6")
login.click
time.sleep(10) # time for manual reCAPTCHA click
submit = driver.find_element(By.CLASS_NAME,"button_zhKS- false inline-block")
submit.click
#'''